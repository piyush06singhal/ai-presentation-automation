import io
import re
import pandas as pd
import openpyxl
from typing import Dict, List, Tuple
from app.services.utils import clean_column_name, parse_currency, parse_percentage, parse_date
from app.services.schema_detector import detect_column_type
from app.schemas.models import DataTypeEnum

def detect_header_row_index(ws: openpyxl.worksheet.worksheet.Worksheet) -> int:
    """Scans the first 10 rows to detect the boundary index of the column header."""
    max_non_empty = 0
    best_row_idx = 0  # 0-indexed for pandas skiprows
    
    # Scan first 10 rows (openpyxl is 1-indexed)
    scan_limit = min(15, ws.max_row)
    for r in range(1, scan_limit + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        non_empty_cnt = sum(1 for x in row_vals if x is not None and str(x).strip() != "")
        
        # We prefer rows with high cell density
        if non_empty_cnt > max_non_empty:
            max_non_empty = non_empty_cnt
            best_row_idx = r - 1  # Convert to 0-indexed for pandas
            
    return best_row_idx

def ingest_workbook(file_bytes: bytes) -> Dict[str, pd.DataFrame]:
    """Loads, discovers worksheets, cleans messy headers, and normalizes cell datatypes."""
    df_collection: Dict[str, pd.DataFrame] = {}
    
    # 1. Load worksheet objects via openpyxl to scan structure
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet_names = wb.sheetnames
    except Exception as e:
        raise ValueError(f"Unable to read Excel workbook structure: {str(e)}")
        
    for name in sheet_names:
        ws = wb[name]
        if ws.max_row <= 1 or ws.max_column == 0:
            # Skip empty worksheets
            continue
            
        # Detect header row index
        skip_rows = detect_header_row_index(ws)
        
        # Read sheet via Pandas
        try:
            df = pd.read_excel(
                io.BytesIO(file_bytes),
                sheet_name=name,
                skiprows=skip_rows
            )
        except Exception:
            continue
            
        if df.empty:
            continue
            
        # Clean headers
        cleaned_columns = []
        seen = {}
        for col in df.columns:
            c_name = clean_column_name(col)
            # Remove pandas-added duplicate suffix e.g. .1, .2
            c_name = re.sub(r'\.\d+$', '', c_name)
            if not c_name:
                c_name = "Unnamed"
            # Resolve duplicates
            if c_name in seen:
                seen[c_name] += 1
                c_name = f"{c_name}_{seen[c_name]}"
            else:
                seen[c_name] = 0
            cleaned_columns.append(c_name)
            
        df.columns = cleaned_columns
        
        # Data cleansing: clean formatting on currency/percentage columns
        for col in df.columns:
            # Infer column type on raw strings
            inferred_type = detect_column_type(col, df[col])
            
            if inferred_type == DataTypeEnum.CURRENCY:
                df[col] = df[col].apply(parse_currency)
            elif inferred_type == DataTypeEnum.PERCENTAGE:
                df[col] = df[col].apply(parse_percentage)
            elif inferred_type == DataTypeEnum.DATE:
                df[col] = df[col].apply(parse_date)
            elif inferred_type == DataTypeEnum.NUMERIC:
                df[col] = pd.to_numeric(df[col], errors='coerce')
                
        # Fill completely null columns or empty fields gracefully
        # Note: Avoid global fillna(0) to prevent distorting categorical counts
        # We fill numeric columns with NaN rather than 0 so stats are calculated correctly
        df_collection[name] = df
        
    return df_collection
