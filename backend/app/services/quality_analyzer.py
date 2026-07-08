import io
import openpyxl
import pandas as pd
from typing import Dict, List
from app.schemas.models import WorkbookHealth, HealthIssue, SeverityEnum

def analyze_workbook_quality(file_bytes: bytes, df_collection: Dict[str, pd.DataFrame]) -> WorkbookHealth:
    """Performs static analysis on the Excel workbook structure and data consistency."""
    issues: List[HealthIssue] = []
    
    # 1. Openpyxl checks (Hidden sheets, merged cells, general sizing)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=False, data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            # Check for empty sheet
            if ws.max_row <= 1 or ws.max_column == 0:
                issues.append(HealthIssue(
                    severity=SeverityEnum.HIGH,
                    warning=f"Worksheet '{name}' is empty.",
                    recommendation="Remove the empty worksheet or populate data before generating.",
                    worksheet=name
                ))
                continue
            # Check for hidden sheet
            if ws.sheet_state in ['hidden', 'veryHidden']:
                issues.append(HealthIssue(
                    severity=SeverityEnum.LOW,
                    warning=f"Worksheet '{name}' is hidden in the workbook.",
                    recommendation="Ensure hidden sheets do not contain critical business metrics needed for the slide deck.",
                    worksheet=name
                ))
            # Check for merged cells
            if len(ws.merged_cells.ranges) > 0:
                issues.append(HealthIssue(
                    severity=SeverityEnum.MEDIUM,
                    warning=f"Worksheet '{name}' contains {len(ws.merged_cells.ranges)} merged cell ranges.",
                    recommendation="Merged cells can distort header extraction. Consider converting to flat columns.",
                    worksheet=name
                ))
            # Check duplicate headers in raw cells
            header_idx = 0
            max_non_empty = 0
            scan_limit = min(15, ws.max_row)
            for r in range(1, scan_limit + 1):
                row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                non_empty_cnt = sum(1 for x in row_vals if x is not None and str(x).strip() != "")
                if non_empty_cnt > max_non_empty:
                    max_non_empty = non_empty_cnt
                    header_idx = r
            if header_idx > 0:
                header_cells = [str(ws.cell(row=header_idx, column=c).value).strip() 
                                for c in range(1, ws.max_column + 1) 
                                if ws.cell(row=header_idx, column=c).value is not None]
                seen_raw = set()
                dup_raw = set()
                for cell_val in header_cells:
                    if cell_val in seen_raw:
                        dup_raw.add(cell_val)
                    seen_raw.add(cell_val)
                if dup_raw:
                    issues.append(HealthIssue(
                        severity=SeverityEnum.MEDIUM,
                        warning=f"Worksheet '{name}' contains duplicate column headers: {list(dup_raw)}.",
                        recommendation="Unique names prevent column overrides. Ensure each column has a distinct name.",
                        worksheet=name
                    ))
    except Exception as e:
        issues.append(HealthIssue(
            severity=SeverityEnum.HIGH,
            warning=f"Failed to perform openpyxl structural checks: {str(e)}",
            recommendation="Confirm the upload file is a valid Excel workbook.",
            worksheet="Workbook"
        ))

    # 2. Pandas-based dataframe inspections
    for sheet_name, df in df_collection.items():
        if df.empty:
            issues.append(HealthIssue(
                severity=SeverityEnum.HIGH,
                warning=f"Worksheet '{sheet_name}' is empty.",
                recommendation="Remove the worksheet or populate data before regenerating.",
                worksheet=sheet_name
            ))
            continue

        # Check duplicate rows
        dup_rows_count = df.duplicated().sum()
        if dup_rows_count > 0:
            issues.append(HealthIssue(
                severity=SeverityEnum.LOW,
                warning=f"Worksheet '{sheet_name}' contains {dup_rows_count} duplicate records.",
                recommendation="Inspect data sources to remove duplicated transactional records.",
                worksheet=sheet_name
            ))

        # Inspect missing values and empty columns
        null_counts = df.isnull().sum()
        empty_cols = []
        high_null_cols = []
        for col_name, null_cnt in null_counts.items():
            if null_cnt == len(df):
                empty_cols.append(col_name)
            elif null_cnt > len(df) * 0.15: # >15% missing
                high_null_cols.append((col_name, int((null_cnt / len(df)) * 100)))

        if empty_cols:
            issues.append(HealthIssue(
                severity=SeverityEnum.MEDIUM,
                warning=f"Worksheet '{sheet_name}' contains columns with 100% missing values: {empty_cols}.",
                recommendation="Remove empty columns to clean up the worksheet.",
                worksheet=sheet_name
            ))
            
        if high_null_cols:
            warnings_str = ", ".join([f"{col} ({pct}% null)" for col, pct in high_null_cols])
            issues.append(HealthIssue(
                severity=SeverityEnum.MEDIUM,
                warning=f"Worksheet '{sheet_name}' contains columns with high missing ratios (>15%): {warnings_str}.",
                recommendation="Ensure these missing fields do not disrupt KPI calculations.",
                worksheet=sheet_name
            ))

        # Check mixed datatypes in columns
        for col in df.columns:
            types = df[col].dropna().apply(lambda x: type(x).__name__).unique()
            if len(types) > 1:
                # If there's a mix of numeric (int, float) and non-numeric (str), flag it
                has_numeric = any('int' in t or 'float' in t for t in types)
                has_str = any('str' in t for t in types)
                if has_numeric and has_str:
                    issues.append(HealthIssue(
                        severity=SeverityEnum.MEDIUM,
                        warning=f"Column '{col}' in worksheet '{sheet_name}' has mixed datatypes: {list(types)}.",
                        recommendation="Clean column values to ensure numeric formulas evaluate correctly.",
                        worksheet=sheet_name,
                        column=str(col)
                    ))

    # Calculate overall health score
    # Scoring: Starts at 100. Subtract based on issue severity:
    # High: -20, Medium: -10, Low: -3
    deductions = 0
    for issue in issues:
        if issue.severity == SeverityEnum.HIGH:
            deductions += 20
        elif issue.severity == SeverityEnum.MEDIUM:
            deductions += 10
        elif issue.severity == SeverityEnum.LOW:
            deductions += 3

    score = max(0.0, min(100.0, 100.0 - deductions))
    
    return WorkbookHealth(issues=issues, overall_score=score)
